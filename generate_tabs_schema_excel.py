import os
import sys

try:
    import openpyxl
except ImportError:
    print("Installing 'openpyxl' package...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_multi_tab_excel():
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    data_font = Font(name="Calibri", size=11)
    border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center")

    headers = ["Field Label (UI)", "Field Name (Code)", "Data Type", "Constraints / Requirements", "Dropdown / Predefined Options", "Description & Behavior"]

    tabs_data = {
        "Incident Form": {
            "title": "Incident Configuration Form Schema & Dropdowns",
            "fields": [
                ["Incident Number", "number", "String", "Read-only, Auto-generated", "None", "Auto-generated identifier for the incident record."],
                ["Created on", "created_on", "Datetime", "Read-only, Auto-generated", "None", "Timestamp when the record was created."],
                ["Status", "status", "Dropdown (String)", "Required (Default: 'new')", 
                 "new (New)\ndiagnosis (Diagnosis)\nin_progress (Work in Progress)\non_hold (On Hold)\nrepair_completed (Repair Completed)\nquality_check (Quality Check)\nclosure (Closure)", 
                 "Current status of the incident lifecycle. Shows 7 different workflow values."],
                ["Priority", "priority", "Dropdown (String)", "Required (Default: 'medium')", 
                 "critical (1 - Critical (AOG))\nhigh (2 - High)\nmedium (3 - Medium)\nlow (4 - Low)", 
                 "SLA priority level of the incident."],
                ["Issue Type", "issue_type", "Dropdown (String)", "Nullable", 
                 "mechanical (Mechanical components)\nsoftware (Software bugs/crashes)\npayload_mission (Payload sensor/mission error)\nperformance (Degraded performance)\nphysical_damage (Hardware physical damage)\nmaintenance_spare (Maintenance request)", 
                 "Classification of the technical issue. Shows 6 system category values."],
                ["SLA", "sla", "String", "Nullable", "None", "Associated service level agreement details."],
                ["Assignment Group", "assignment_group", "String / Reference", "Nullable", "None", "The group/team assigned to resolve this ticket."],
                ["Assigned to", "assigned_to", "String / Reference", "Nullable", "None", "The specific agent/technician assigned to resolve this ticket."],
                ["Customer Name", "caller", "Dropdown (String)", "Required", 
                 "Acme Rail Systems\nMetroLine Services\nUrban Transit Corp", 
                 "Selecting a customer filters the customer contracts, requestor fields, and platform variants in the UI."],
                ["Customer Contract", "customer_contract", "Dropdown (String)", "Nullable", 
                 "CTR-2024-001 (Filters on customer 'Acme Rail Systems')\nCTR-2024-014 (Filters on customer 'MetroLine Services')\nCTR-2023-220 (Filters on customer 'Urban Transit Corp')", 
                 "Selectable active contracts. Filtered dynamically based on the selected Customer Name."],
                ["Requestor Name", "requestor_name", "Dropdown / Datalist (String)", "Nullable", 
                 "Ravi Kumar (Acme)\nNeha Singh (MetroLine)\nAmit Rao (MetroLine)\nArjun Mehta (Urban Transit)\nKavya Nair (Urban Transit)\nRahul Iyer (Urban Transit)", 
                 "Name of the customer requestor reporting the incident. Auto-populates contact when selected."],
                ["Requestor Contact", "requestor_contact", "Dropdown / Datalist (String)", "Nullable", 
                 "+91 98765 43210 (Ravi Kumar)\n+91 91234 56789 (Neha Singh)\n+91 90000 11223 (Amit Rao)\n+91 99887 66554 (Arjun Mehta)\n+91 91111 22334 (Kavya Nair)\n+91 93333 44556 (Rahul Iyer)", 
                 "Contact number for the requestor. Automatically synced based on Requestor Name selection."],
                ["SRLM System", "srlm_system", "Dropdown (String)", "Nullable", 
                 "LM (Loitering Munition)\nGCS (Ground Control System)\nTMV (Tactical Mobility Vehicle)\nSimulator (Simulator)", 
                 "Selects system class. Selecting this filters Platform Variant and Sub-system options."],
                ["Platform Variant", "platform_variant", "Dropdown (String)", "Nullable", 
                 "LM Alpha - (LM-0001)\nLM Bravo - (LM-0002)\nGCS North - (GCS-0101)\nTMV Ranger - (TMV-2201)\nSimulator Prime - (SIM-3001)", 
                 "List of system platform variants. Filtered dynamically based on Customer, Contract, and SRLM System."],
                ["Sub-system", "sub_system", "Dropdown (String)", "Nullable", 
                 "For LM:\n- Airframe\n- Propulsion\n- Avionics & Flight Control\n- Communication\n- Payload\n\nFor GCS:\n- Computing\n- Communication\n- Display\n- Control Interface\n- Power\n\nFor TMV:\n- Powertrain\n- Electrical\n- Communication\n- Navigation\n- Payload Handling\n\nFor Simulator:\n- Computing\n- Display\n- Control\n- Audio\n- Software", 
                 "Selectable sub-systems. Filtered dynamically based on the selected SRLM System."],
                ["Line Replaceable Unit (LRU)", "line_replaceable_unit", "Dropdown (String)", "Nullable", 
                 "LM Alpha / Navigation:\n- Guidance Module (LRU-1001)\nLM Alpha / Airframe:\n- Airframe Sensor Pack (LRU-1002)\nLM Alpha / Propulsion:\n- Propulsion Pump (LRU-1004)\nGCS North / Computing:\n- Compute Node A (LRU-2101)\nGCS North / Communication:\n- Comms Router (LRU-2103)\nGCS North / Display:\n- Display Panel A (LRU-2105)\nTMV Ranger / Powertrain:\n- Transmission Module (LRU-3101)\nTMV Ranger / Electrical:\n- Battery Pack (LRU-3103)\nTMV Ranger / Navigation:\n- GPS Module (LRU-3108)\nSimulator Prime / Computing:\n- Sim Compute Node (LRU-4101)\nSimulator Prime / Display:\n- Projection Unit (LRU-4104)\nSimulator Prime / Audio:\n- Audio Mixer (LRU-4107)", 
                 "Specific component part (LRU) being logged. Filtered dynamically based on SRLM System, Platform Variant, and Sub-system."],
                ["Warranty Status", "warranty_status", "String", "Read-only / Display", "None", "Displays warranty status dynamically based on selected product variant (e.g. Active, Expired)."],
                ["Last Serviced on", "last_serviced_on", "String", "Read-only / Display", "None", "Displays date of last maintenance action dynamically."],
                ["Short description", "title", "String", "Required", "None", "A concise overview of the incident."],
                ["Description", "description", "Text", "Nullable", "None", "Detailed logs or explanation of the incident."]
            ]
        },
        "Contracts": {
            "title": "Contract Configuration Form Schema",
            "fields": [
                ["Contract Number", "contract_number", "String", "Required", "None", "Unique identification number of the contract."],
                ["Customer", "customer", "Dropdown (String)", "Required", "Acme Rail Systems\nMetroLine Services\nUrban Transit Corp", "Dropdown linking the contract to an existing customer record."],
                ["Executed on", "executed_on", "Date / String", "Required", "None", "Execution date of the contract (YYYY-MM-DD)."],
                ["Valid Till", "valid_till", "Date / String", "Required", "None", "Expiration date of the contract (YYYY-MM-DD)."],
                ["Status", "status", "Dropdown (String)", "Required", "Active\nExpired\nDisabled", "Lifecycle status of the contract."],
                ["Attachment", "attachment", "File", "Nullable", "None", "Upload field for supporting contract PDF/Doc documents."],
                ["No of Deliverables [LM/ GCS/ TMV/ RDV/ Simulator/ LRU/ MRLS/ SMTs/STEs]", "main_deliverables", "String", "Nullable", "None", "No of Deliverables (LM/ GCS/ TMV/ RDV/ Simulator/ LRU/ MRLS/ SMTs/STEs)."],
                ["Includes warranty as per executed on and valid till", "main_warranty", "Boolean", "Nullable (Default: False)", "None", "Includes warranty as per executed on and valid till."],
                ["Manuals [All versions with date of issue]", "main_manuals", "String", "Nullable", "None", "Manuals (All versions with date of issue)."],
                ["Scheduled Maintenance [Periodic Maintenance] (option to include spares for replacement)", "sub_scheduled_maintenance", "Boolean", "Nullable (Default: False)", "None", "Scheduled Maintenance (Periodic Maintenance) (option to include spares for replacement)."],
                ["Unscheduled Maintenance [Repair of items post failure]", "sub_unscheduled_maintenance", "Boolean", "Nullable (Default: False)", "None", "Unscheduled Maintenance (Repair of items post failure)."],
                ["Calibration of equipments", "sub_calibration", "Boolean", "Nullable (Default: False)", "None", "Calibration of equipments."],
                ["Software upgrade", "sub_software_upgrade", "Boolean", "Nullable (Default: False)", "None", "Software upgrade."],
                ["Record of No of scheduled/ unscheduled visits (including no of personnel and days)", "sub_visits_record", "String", "Nullable", "None", "Record of No of scheduled/ unscheduled visits (including no of personnel and days)."],
                ["Refresher Training", "sub_refresher_training", "Boolean", "Nullable (Default: False)", "None", "Refresher Training."]
            ]
        },
        "Customers": {
            "title": "Customer Configuration Form Schema",
            "fields": [
                ["Customer name", "customer_name", "String", "Required", "None", "Official company or client name."],
                ["Primary address", "primary_address", "String", "Required", "None", "Main billing/shipping location address."],
                ["Contact name", "primary_contact_name", "String", "Required", "None", "Full name of primary customer contact personnel."],
                ["Contact email", "primary_contact_email", "String", "Required", "None", "Email address of primary customer contact."],
                ["Contact phone", "primary_contact_phone", "String", "Required", "None", "Phone number of primary customer contact."],
                ["Additional contact name", "contact_name[]", "String", "Nullable", "None", "List of secondary/additional contact names."],
                ["Additional contact email", "contact_email[]", "String", "Nullable", "None", "List of secondary/additional contact emails."],
                ["Additional contact phone", "contact_phone[]", "String", "Nullable", "None", "List of secondary/additional contact phone numbers."]
            ]
        },
        "Loitering Munition": {
            "title": "Loitering Munition (LM) Unit Configuration Form Schema",
            "fields": [
                ["Serial Number", "serial_number", "String", "Required", "None", "Unique military grade hardware serial number."],
                ["Unit Name", "unit_name", "String", "Required", "None", "Descriptive name for the individual unit (e.g. LM Alpha)."],
                ["SRLM System", "srlm_system", "Dropdown (String)", "Read-only (Default: 'LM')", "LM (Loitering Munition)", "System categorization, fixed to LM."],
                ["Customer", "customer_id", "Dropdown (String)", "Required", "Acme Rail Systems\nMetroLine Services\nUrban Transit Corp", "Assigns ownership of the unit to a customer."],
                ["Contract", "contract_id", "Dropdown (String)", "Required", "CTR-2024-001\nCTR-2024-014\nCTR-2023-220", "Links the unit deployment to an active contract. Filtered by selected Customer."],
                ["Warranty Valid From", "warranty_valid_from", "Date / String", "Required", "None", "Warranty coverage start date (YYYY-MM-DD)."],
                ["Warranty Valid To", "warranty_valid_to", "Date / String", "Required", "None", "Warranty coverage end date (YYYY-MM-DD)."]
            ]
        },
        "Batteries": {
            "title": "Batteries Configuration Form Schema",
            "fields": [
                ["SAP Part number", "sap_part_number", "String", "Required, Length: 14", "None", "14 Digit part no generated on SAP."],
                ["Customer Part no", "customer_part_number", "String", "Required, Length: 8", "None", "08 Digit part no generated for customer."],
                ["Serial Number", "serial_number", "String", "Required", "None", "Unique hardware serial number."],
                ["Unit Name", "unit_name", "String", "Required", "None", "Descriptive name for the individual unit (e.g. Battery Pack)."],
                ["Battery Type", "battery_type", "Dropdown (String)", "Required", "Lithium-Ion 12S 44.4V\nLi-Polymer 6S 22.2V\nLi-Fe 4S 12.8V", "Battery type must include the capacity and cells."],
                ["Customer", "customer_id", "Dropdown (String)", "Required", "Acme Rail Systems\nMetroLine Services\nUrban Transit Corp", "Assigns ownership of the unit to a customer."],
                ["Contract", "contract_id", "Dropdown (String)", "Required", "CTR-2024-001\nCTR-2024-014\nCTR-2023-220", "Links the unit deployment to an active contract. Filtered by selected Customer."],
                ["Warranty Valid From", "warranty_valid_from", "Date / String", "Required", "None", "Warranty coverage start date (YYYY-MM-DD)."],
                ["Warranty Valid To", "warranty_valid_to", "Date / String", "Required", "None", "Warranty coverage end date (YYYY-MM-DD)."],
                ["Battery and its types", "battery_and_its_types", "String", "Required", "None", "Details about battery composition and capacity."],
                ["Battery Chargers", "battery_chargers", "String", "Required", "None", "Supported battery charger models."],
                ["Make", "make", "String", "Required", "None", "Manufacturer make."],
                ["Model", "model", "String", "Required", "None", "Manufacturer model."],
                ["Software version", "software_version", "String", "Required", "None", "Installed software/firmware version."]
            ]
        },
        "War Head": {
            "title": "War Head Unit Configuration Form Schema",
            "fields": [
                ["SAP Part number", "sap_part_number", "String", "Required, Length: 14", "None", "14 Digit part no generated on SAP."],
                ["Customer Part no", "customer_part_number", "String", "Required, Length: 8", "None", "08 Digit part no generated for customer."],
                ["Serial Number", "serial_number", "String", "Required", "None", "Unique military grade hardware serial number."],
                ["Unit Name", "unit_name", "String", "Required", "None", "Descriptive name for the individual unit (e.g. War Head)."],
                ["SRLM System", "srlm_system", "Dropdown (String)", "Read-only (Default: 'LM')", "LM (Loitering Munition)", "System categorization, fixed to LM."],
                ["Customer", "customer_id", "Dropdown (String)", "Required", "Acme Rail Systems\nMetroLine Services\nUrban Transit Corp", "Assigns ownership of the unit to a customer."],
                ["Contract", "contract_id", "Dropdown (String)", "Required", "CTR-2024-001\nCTR-2024-014\nCTR-2023-220", "Links the unit deployment to an active contract. Filtered by selected Customer."],
                ["Warranty Valid From", "warranty_valid_from", "Date / String", "Required", "None", "Warranty coverage start date (YYYY-MM-DD)."],
                ["Warranty Valid To", "warranty_valid_to", "Date / String", "Required", "None", "Warranty coverage end date (YYYY-MM-DD)."],
                ["Make", "make", "String", "Required", "None", "Manufacturer make."],
                ["Model", "model", "String", "Required", "None", "Manufacturer model."],
                ["Software version", "software_version", "String", "Required", "None", "Installed software version."]
            ]
        },
        "Ground Control System": {
            "title": "Ground Control System (GCS) Unit Configuration Form Schema",
            "fields": [
                ["SAP Part number", "sap_part_number", "String", "Required, Length: 14", "None", "14 Digit part no generated on SAP."],
                ["Customer Part no", "customer_part_number", "String", "Required, Length: 8", "None", "08 Digit part no generated for customer."],
                ["Serial Number", "serial_number", "String", "Required", "None", "Unique system hardware serial number."],
                ["Unit Name", "unit_name", "String", "Required", "None", "Descriptive unit name (e.g. GCS North)."],
                ["SRLM System", "srlm_system", "Dropdown (String)", "Read-only (Default: 'GCS')", "GCS (Ground Control System)", "System categorization, fixed to GCS."],
                ["Customer", "customer_id", "Dropdown (String)", "Required", "Acme Rail Systems\nMetroLine Services\nUrban Transit Corp", "Assigns ownership of the unit to a customer."],
                ["Contract", "contract_id", "Dropdown (String)", "Required", "CTR-2024-001\nCTR-2024-014\nCTR-2023-220", "Links the unit deployment to an active contract. Filtered by selected Customer."],
                ["Warranty Valid From", "warranty_valid_from", "Date / String", "Required", "None", "Warranty coverage start date (YYYY-MM-DD)."],
                ["Warranty Valid To", "warranty_valid_to", "Date / String", "Required", "None", "Warranty coverage end date (YYYY-MM-DD)."],
                ["Make", "make", "String", "Required", "None", "Manufacturer make."],
                ["Model", "model", "String", "Required", "None", "Manufacturer model."],
                ["Software version", "software_version", "String", "Required", "None", "Installed software version."]
            ]
        },
        "Tactical Mobility Vehicle": {
            "title": "Tactical Mobility Vehicle (TMV) Unit Configuration Form Schema",
            "fields": [
                ["SAP Part number", "sap_part_number", "String", "Required, Length: 14", "None", "14 Digit part no generated on SAP."],
                ["Customer Part no", "customer_part_number", "String", "Required, Length: 8", "None", "08 Digit part no generated for customer."],
                ["Serial Number", "serial_number", "String", "Required", "None", "Unique hardware chassis/VIN code."],
                ["Unit Name", "unit_name", "String", "Required", "None", "Descriptive unit name (e.g. TMV Ranger)."],
                ["SRLM System", "srlm_system", "Dropdown (String)", "Read-only (Default: 'TMV')", "TMV (Tactical Mobility Vehicle)", "System categorization, fixed to TMV."],
                ["Customer", "customer_id", "Dropdown (String)", "Required", "Acme Rail Systems\nMetroLine Services\nUrban Transit Corp", "Assigns ownership of the vehicle to a customer."],
                ["Contract", "contract_id", "Dropdown (String)", "Required", "CTR-2024-001\nCTR-2024-014\nCTR-2023-220", "Links the unit deployment to an active contract. Filtered by selected Customer."],
                ["Warranty Valid From", "warranty_valid_from", "Date / String", "Required", "None", "Warranty coverage start date (YYYY-MM-DD)."],
                ["Warranty Valid To", "warranty_valid_to", "Date / String", "Required", "None", "Warranty coverage end date (YYYY-MM-DD)."],
                ["Make", "make", "String", "Nullable", "None", "Manufacturer of the vehicle."],
                ["Model", "model", "String", "Nullable", "None", "Model details of the vehicle."]
            ]
        },
        "Simulator": {
            "title": "Simulator Unit Configuration Form Schema",
            "fields": [
                ["Serial Number", "serial_number", "String", "Required", "None", "Unique system hardware serial number."],
                ["Unit Name", "unit_name", "String", "Required", "None", "Descriptive unit name (e.g. Simulator Prime)."],
                ["SRLM System", "srlm_system", "Dropdown (String)", "Read-only (Default: 'Simulator')", "Simulator", "System categorization, fixed to Simulator."],
                ["Customer", "customer_id", "Dropdown (String)", "Required", "Acme Rail Systems\nMetroLine Services\nUrban Transit Corp", "Assigns ownership of the simulator to a customer."],
                ["Contract", "contract_id", "Dropdown (String)", "Required", "CTR-2024-001\nCTR-2024-014\nCTR-2023-220", "Links the unit deployment to an active contract. Filtered by selected Customer."],
                ["Warranty Valid From", "warranty_valid_from", "Date / String", "Required", "None", "Warranty coverage start date (YYYY-MM-DD)."],
                ["Warranty Valid To", "warranty_valid_to", "Date / String", "Required", "None", "Warranty coverage end date (YYYY-MM-DD)."]
            ]
        },
        "LRU": {
            "title": "Line Replaceable Unit (LRU) Configuration Form Schema",
            "fields": [
                ["LRU Serial Number", "lru_serial_number", "String", "Required", "None", "Unique serial number of the LRU part."],
                ["LRU Name", "lru_name", "String", "Required", "None", "Descriptive component part name (e.g. Flight Computer)."],
                ["SAP Part Number", "sap_part_number", "String", "Required, Length: 14", "None", "14 digit part no generated on SAP."],
                ["Customer Part Number", "customer_part_number", "String", "Required, Length: 8", "None", "08 digit part no generated for customer."],
                ["Make", "make", "String", "Nullable", "None", "Manufacturer of the LRU part."],
                ["Model", "model", "String", "Nullable", "None", "Model number of the LRU part."],
                ["Software Version", "software_version", "String", "Nullable", "None", "Installed firmware/software version on the LRU."],
                ["Related Platform Variant Name", "platform_variant", "Dropdown (String)", "Required", 
                 "LM Alpha - (LM-0001)\nLM Bravo - (LM-0002)\nGCS North - (GCS-0101)\nTMV Ranger - (TMV-2201)\nSimulator Prime - (SIM-3001)", 
                 "Links this specific LRU assembly to a parent platform variant type."],
                ["Sub-system", "sub_system", "String", "Required", "None", "Associated structural sub-system (e.g. Propulsion, Computing, Electrical)."]
            ]
        },
        "SAM": {
            "title": "Surface-to-Air Missile (SAM) Unit Configuration Form Schema",
            "fields": [
                ["SAP Part number", "sap_part_number", "String", "Required, Length: 14", "None", "14 Digit part no generated on SAP."],
                ["Customer Part no", "customer_part_number", "String", "Required, Length: 8", "None", "08 Digit part no generated for customer."],
                ["Serial Number", "serial_number", "String", "Required", "None", "Unique hardware serial number."],
                ["Unit Name", "unit_name", "String", "Required", "None", "Descriptive name for the individual unit (e.g. SAM System Alpha)."],
                ["SRLM System", "srlm_system", "Dropdown (String)", "Read-only (Default: 'LM')", "LM (Loitering Munition)", "System categorization, fixed to LM."],
                ["Customer", "customer_id", "Dropdown (String)", "Required", "Acme Rail Systems\nMetroLine Services\nUrban Transit Corp", "Assigns ownership of the unit to a customer."],
                ["Contract", "contract_id", "Dropdown (String)", "Required", "CTR-2024-001\nCTR-2024-014\nCTR-2023-220", "Links the unit deployment to an active contract. Filtered by selected Customer."],
                ["Warranty Valid From", "warranty_valid_from", "Date / String", "Required", "None", "Warranty coverage start date (YYYY-MM-DD)."],
                ["Warranty Valid To", "warranty_valid_to", "Date / String", "Required", "None", "Warranty coverage end date (YYYY-MM-DD)."],
                ["Make", "make", "String", "Required", "None", "Manufacturer make."],
                ["Model", "model", "String", "Required", "None", "Manufacturer model."],
                ["Software version", "software_version", "String", "Required", "None", "Installed software version."]
            ]
        },
        "MRLS": {
            "title": "Manufacturers Recommended List of Spares (MRLS) Configuration Form Schema",
            "fields": [
                ["SAP Part number", "sap_part_number", "String", "Required, Length: 14", "None", "14 Digit part no generated on SAP."],
                ["Customer Part no", "customer_part_number", "String", "Required, Length: 8", "None", "08 Digit part no generated for customer."],
                ["Serial Number", "serial_number", "String", "Required", "None", "Unique spare kit serial number."],
                ["Unit Name", "unit_name", "String", "Required", "None", "Descriptive name for the individual unit (e.g. MRLS Spare Kit)."],
                ["Related Platform Variant Name", "platform_variant", "Dropdown (String)", "Required", 
                 "LM Alpha - (LM-0001)\nLM Bravo - (LM-0002)\nGCS North - (GCS-0101)\nTMV Ranger - (TMV-2201)\nSimulator Prime - (SIM-3001)", 
                 "Links this spare kit assembly to a parent platform variant type."],
                ["Sub-system", "sub_system", "String", "Required", "None", "Associated structural sub-system."],
                ["LRU Serial Number", "lru_serial_number", "String", "Required", "None", "Associated LRU serial number."],
                ["LRU Name", "lru_name", "String", "Required", "None", "Associated LRU name."],
                ["Customer", "customer_id", "Dropdown (String)", "Required", "Acme Rail Systems\nMetroLine Services\nUrban Transit Corp", "Assigns ownership of the unit to a customer."],
                ["Contract", "contract_id", "Dropdown (String)", "Required", "CTR-2024-001\nCTR-2024-014\nCTR-2023-220", "Links the unit deployment to an active contract. Filtered by selected Customer."],
                ["Warranty Valid From", "warranty_valid_from", "Date / String", "Required", "None", "Warranty coverage start date (YYYY-MM-DD)."],
                ["Warranty Valid To", "warranty_valid_to", "Date / String", "Required", "None", "Warranty coverage end date (YYYY-MM-DD)."],
                ["Make", "make", "String", "Required", "None", "Manufacturer make."],
                ["Model", "model", "String", "Required", "None", "Manufacturer model."],
                ["Software version", "software_version", "String", "Required", "None", "Installed software version."],
                ["Spares Table: Nomenclature", "spare_nomenclature[]", "Table Column (String)", "Required for table row", "None", "Name/Nomenclature of the spare item in the spares table."],
                ["Spares Table: Part No.", "spare_part_no[]", "Table Column (String)", "Required for table row", "None", "Part number of the spare item in the spares table."],
                ["Spares Table: Quantity Installed", "spare_qty_installed[]", "Table Column (Integer)", "Required for table row", "None", "Quantity installed in the system."],
                ["Spares Table: Numbers Provided in Spares", "spare_qty_provided[]", "Table Column (Integer)", "Required for table row", "None", "Numbers of spares provided."]
            ]
        },
        "SMT STE": {
            "title": "Field Assembly Tools / SMT / STE Configuration Form Schema",
            "fields": [
                ["SAP Part number", "sap_part_number", "String", "Required, Length: 14", "None", "14 Digit part no generated on SAP."],
                ["Customer Part no", "customer_part_number", "String", "Required, Length: 8", "None", "08 Digit part no generated for customer."],
                ["Serial Number", "serial_number", "String", "Required", "None", "Unique tools kit serial number."],
                ["Unit Name", "unit_name", "String", "Required", "None", "Descriptive name for the toolset unit (e.g. SMT-STE Toolset)."],
                ["Related Platform Variant Name", "platform_variant", "Dropdown (String)", "Required", 
                 "LM Alpha - (LM-0001)\nLM Bravo - (LM-0002)\nGCS North - (GCS-0101)\nTMV Ranger - (TMV-2201)\nSimulator Prime - (SIM-3001)", 
                 "Links this toolset to a parent platform variant type."],
                ["Sub-system", "sub_system", "String", "Required", "None", "Associated structural sub-system."],
                ["LRU Serial Number", "lru_serial_number", "String", "Required", "None", "Associated LRU serial number."],
                ["LRU Name", "lru_name", "String", "Required", "None", "Associated LRU name."],
                ["Customer", "customer_id", "Dropdown (String)", "Required", "Acme Rail Systems\nMetroLine Services\nUrban Transit Corp", "Assigns ownership of the unit to a customer."],
                ["Contract", "contract_id", "Dropdown (String)", "Required", "CTR-2024-001\nCTR-2024-014\nCTR-2023-220", "Links the unit deployment to an active contract. Filtered by selected Customer."],
                ["Warranty Valid From", "warranty_valid_from", "Date / String", "Required", "None", "Warranty coverage start date (YYYY-MM-DD)."],
                ["Warranty Valid To", "warranty_valid_to", "Date / String", "Required", "None", "Warranty coverage end date (YYYY-MM-DD)."],
                ["Make", "make", "String", "Required", "None", "Manufacturer make."],
                ["Model", "model", "String", "Required", "None", "Manufacturer model."],
                ["Software version", "software_version", "String", "Required", "None", "Installed software version."],
                ["Tools Table: Component/LRU", "tool_component[]", "Table Column (String)", "Required for table row", "None", "Associated Component/LRU for the tool in the tools table."],
                ["Tools Table: Part No.", "tool_part_no[]", "Table Column (String)", "Required for table row", "None", "Part number of the tool in the tools table."],
                ["Tools Table: Qty", "tool_qty[]", "Table Column (Integer)", "Required for table row", "None", "Quantity of tools provided."]
            ]
        }
    }

    for tab_name, tab_info in tabs_data.items():
        ws = wb.create_sheet(title=tab_name)
        ws.views.sheetView[0].showGridLines = True

        # Write Title
        ws.cell(row=2, column=1, value=tab_info["title"]).font = title_font
        ws.row_dimensions[2].height = 25

        # Write Headers
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[4].height = 24

        # Write Data
        current_row = 5
        for field in tab_info["fields"]:
            # Adjust row height if options have newlines
            options_text = field[4]
            newline_count = options_text.count("\n")
            if newline_count > 0:
                ws.row_dimensions[current_row].height = max(18 * (newline_count + 1), 60)
            else:
                ws.row_dimensions[current_row].height = 24

            for col_idx, val in enumerate(field, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = align_left

            current_row += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[3:]:  # Start checking length from row 4
                if cell.value is not None:
                    lines = str(cell.value).split("\n")
                    for line in lines:
                        max_len = max(max_len, len(line))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        # Apply specific wrap and width settings to Dropdowns & Description columns
        ws.column_dimensions['E'].width = 50  # Dropdown options
        ws.column_dimensions['F'].width = 50  # Description & Behavior

    filename = "Vajra_Forms_Data_Dictionary.xlsx"
    wb.save(filename)
    print(f"[SUCCESS] Multi-tab Excel file generated at: {filename}")

if __name__ == "__main__":
    generate_multi_tab_excel()
