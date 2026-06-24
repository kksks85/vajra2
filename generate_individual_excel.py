import os
import sys
import sqlite3

# Add the root directory to path to import local modules
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
except ImportError:
    print("Installing 'openpyxl' package...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from routers.admin import (
        _get_sample_customers,
        _get_sample_contracts,
        _get_sample_lrus,
    )
except ImportError:
    _get_sample_customers = None

def save_individual_excel(filename, title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.views.sheetView[0].showGridLines = True
    
    # Fonts and styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    data_font = Font(name="Calibri", size=11)
    
    border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Title row (Row 2)
    ws.cell(row=2, column=1, value=title).font = title_font
    ws.row_dimensions[2].height = 25
    
    # Headers row (Row 4)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[4].height = 24
    
    # Data rows (Row 5 onwards)
    for r_idx, r in enumerate(rows, 5):
        ws.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = align_left
            
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Max length starting from row 4
        for cell in col[3:]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(filename)
    print(f"[SUCCESS] Saved Excel sheet to: {filename}")

def main():
    # 1. Generate incidents.xlsx (ID, title, description, status, priority, stage, dates)
    inc_headers = ["ID", "Title", "Description", "Status", "Priority", "Stage", "Created At", "Updated At"]
    inc_rows = []
    
    # Query db first
    db_file = "vajra.db"
    db_incidents = []
    if os.path.exists(db_file):
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT id, title, description, status, priority, stage, created_at, updated_at FROM incidents")
            db_incidents = cursor.fetchall()
        except sqlite3.OperationalError:
            pass
        conn.close()
        
    if db_incidents:
        for i in db_incidents:
            status_map = {"new": "New", "diagnosis": "Diagnosis", "in_progress": "Work in Progress", "on_hold": "On Hold", "repair_completed": "Repair Completed", "quality_check": "Quality Check", "closure": "Closure"}
            priority_map = {"critical": "1 - Critical (AOG)", "high": "2 - High", "medium": "3 - Medium", "low": "4 - Low"}
            inc_rows.append([
                f"INC{i[0]:07d}",
                i[1],
                i[2] or "",
                status_map.get(i[3], i[3].capitalize()),
                priority_map.get(i[4], i[4].capitalize()),
                i[5].capitalize() if i[5] else "",
                i[6],
                i[7] or ""
            ])
    else:
        # Prepopulate with 3 sample incidents so it's not empty
        inc_rows = [
            ["INC0000001", "Propulsion system failure during hover test", "Sudden loss of thrust on motor 3. Telemetry indicates motor controller overheating.", "Work in Progress", "1 - Critical (AOG)", "Diagnosis", "2026-06-01 10:00:00", "2026-06-02 14:30:00"],
            ["INC0000002", "GPS signal loss and autopilot fallback", "GPS lock lost for 15 seconds during automated route. Autopilot successfully switched to inertial navigation.", "New", "2 - High", "Triage", "2026-06-03 09:15:00", "2026-06-03 09:15:00"],
            ["INC0000003", "Display console screen flicker on GCS", "Primary display console screen A flickers intermittently during GCS operation.", "Repair Completed", "3 - Medium", "Quality Check", "2026-06-04 11:20:00", "2026-06-06 16:45:00"]
        ]
        
    save_individual_excel("incidents.xlsx", "Service Incidents Log", inc_headers, inc_rows)

    # 2. Generate customers_list.xlsx (Customer ID, Name, Contact, Address, Status)
    cust_headers = ["Customer ID", "Name", "Contact Name", "Contact Email", "Contact Phone", "Address", "Status"]
    cust_rows = []
    if _get_sample_customers:
        for c in _get_sample_customers():
            cust_rows.append([
                c["id"],
                c["name"],
                c["contact_name"],
                c["contact_email"],
                c["contact_phone"],
                c["primary_address"],
                c["status"]
            ])
    save_individual_excel("customers_list.xlsx", "Customer Configuration List", cust_headers, cust_rows)

    # 3. Generate contracts_list.xlsx (Contract ID, Customer Name, Execution/Validity dates, Deliverables, Warranty, Manuals, AMC/CMC details)
    contract_headers = [
        "Contract ID", "Customer Name", "Execution Date", "Validity Date", "Status",
        "No of Deliverables", "Includes Warranty", "Manuals", 
        "Scheduled Maintenance", "Unscheduled Maintenance", 
        "Calibration of Equipments", "Software Upgrade", 
        "Record of Visits", "Refresher Training"
    ]
    contract_rows = []
    if _get_sample_contracts:
        for c in _get_sample_contracts():
            contract_rows.append([
                c["number"], # Using contract number as Contract ID
                c["customer_name"],
                c["executed_on"],
                c["valid_till"],
                c["status"],
                c.get("main_deliverables", ""),
                "Yes" if c.get("main_warranty") else "No",
                c.get("main_manuals", ""),
                "Yes" if c.get("sub_scheduled_maintenance") else "No",
                "Yes" if c.get("sub_unscheduled_maintenance") else "No",
                "Yes" if c.get("sub_calibration") else "No",
                "Yes" if c.get("sub_software_upgrade") else "No",
                c.get("sub_visits_record", ""),
                "Yes" if c.get("sub_refresher_training") else "No"
            ])
    save_individual_excel("contracts_list.xlsx", "Contracts List", contract_headers, contract_rows)

    # 4. Generate lrus.xlsx (LRU ID, Serial, Name, SAP Part Number, Customer Part Number, Make, Model, Software Version, Subsystem, Platform)
    lru_headers = ["LRU ID", "Serial Number", "LRU Name", "SAP Part Number", "Customer Part Number", "Make", "Model", "Software Version", "Subsystem", "Platform Variant", "Status"]
    lru_rows = []
    if _get_sample_lrus:
        for l in _get_sample_lrus():
            lru_rows.append([
                l["id"],
                l["serial_number"],
                l["name"],
                l.get("sap_part_number", ""),
                l.get("customer_part_number", ""),
                l.get("make", ""),
                l.get("model", ""),
                l.get("software_version", ""),
                l["sub_system"],
                l["platform_variant"],
                l["status"]
            ])
    save_individual_excel("lrus.xlsx", "Line Replaceable Units (LRU) List", lru_headers, lru_rows)

if __name__ == "__main__":
    main()
