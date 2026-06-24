import os
import sys

try:
    import openpyxl
except ImportError:
    print("Installing 'openpyxl' package...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_schema_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Database Schema Dictionary"
    ws.views.sheetView[0].showGridLines = True

    # Styling helpers
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Table headers colors to separate tables visually
    table_fill_color = "E9EDF4"
    table_fill = PatternFill(start_color=table_fill_color, end_color=table_fill_color, fill_type="solid")
    table_font = Font(name="Calibri", size=11, bold=True, color="1F497D")
    
    data_font = Font(name="Calibri", size=11)
    border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # Title block
    ws.cell(row=2, column=1, value="Vajra Database Schema & Data Dictionary").font = title_font
    ws.row_dimensions[2].height = 25
    
    # Table headers (Row 4)
    headers = ["Table Name", "Field Name", "Data Type", "Key / Constraints", "Default Value", "Predefined / Dropdown Options"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[4].height = 24

    # Combined Schema Data
    schema_data = [
        # --- incidents ---
        {
            "table": "incidents",
            "fields": [
                ["id", "INTEGER", "Primary Key, Auto-increment", "None", "Unique identifier for each incident report."],
                ["title", "VARCHAR(200)", "Required", "None", "Short description/summary of the incident."],
                ["description", "TEXT", "Nullable", "None", "Detailed description of the issue or error encountered."],
                ["status", "VARCHAR(50)", "Required", "'new'", "Current status of the ticket.\nDropdown Options:\n- 'new' (New)\n- 'diagnosis' (Diagnosis)\n- 'in_progress' (Work in Progress)\n- 'on_hold' (On Hold)\n- 'repair_completed' (Repair Completed)\n- 'quality_check' (Quality Check)\n- 'closure' (Closure)"],
                ["priority", "VARCHAR(50)", "Required", "'medium'", "SLA/Urgency category.\nDropdown Options:\n- 'critical' (1 - Critical (AOG))\n- 'high' (2 - High)\n- 'medium' (3 - Medium)\n- 'low' (4 - Low)"],
                ["issue_type", "VARCHAR(100)", "Nullable", "None", "System category causing the issue.\nDropdown Options:\n- 'mechanical' (Mechanical components)\n- 'software' (Software bugs/crashes)\n- 'payload_mission' (Payload payload sensor/mission error)\n- 'performance' (Degraded performance)\n- 'physical_damage' (Hardware physical damage)\n- 'maintenance_spare' (Maintenance request)"],
                ["stage", "VARCHAR(50)", "Required", "'triage'", "Current workflow builder stage.\nDropdown Options:\n- 'triage'\n- 'diagnosis'\n- 'repair'\n- 'test'\n- 'closure'"],
                ["created_at", "DATETIME", "Required", "func.now()", "Timestamp when the incident was logged."],
                ["updated_at", "DATETIME", "Nullable", "None", "Timestamp when the incident was last modified."]
            ]
        },
        # --- users ---
        {
            "table": "users",
            "fields": [
                ["id", "INTEGER", "Primary Key, Auto-increment", "None", "Unique user identifier."],
                ["username", "VARCHAR(150)", "Required, Unique", "None", "Login username name."],
                ["email", "VARCHAR(200)", "Required, Unique", "None", "User communication email address."],
                ["full_name", "VARCHAR(200)", "Required", "''", "User's full name."],
                ["password_hash", "VARCHAR(255)", "Required", "None", "Bcrypt-hashed password string."],
                ["role", "VARCHAR(50)", "Required", "'agent'", "System privilege role.\nDropdown Options:\n- 'admin' (Administrator)\n- 'supervisor' (Manager/Supervisor)\n- 'technician' (Technician/Engineer)\n- 'agent' (Support agent)"],
                ["license_type", "VARCHAR(50)", "Required", "'Standard'", "Subscription license level.\nDropdown Options:\n- 'Standard'\n- 'Enterprise'"],
                ["department", "VARCHAR(100)", "Required", "''", "Internal organization department.\nDropdown Options:\n- 'Administration'\n- 'Project Management'\n- 'Operations'\n- 'Quality Assurance'\n- 'Field Operations'\n- 'Customer Support'\n- 'Technical Consulting'\n- 'Engineering'"],
                ["location", "VARCHAR(100)", "Required", "''", "Geographic work center.\nDropdown Options:\n- 'Pune HQ'\n- 'Mumbai Branch'\n- 'Bangalore Branch'\n- 'Delhi Branch'\n- 'Hyderabad Branch'"],
                ["employee_id", "VARCHAR(50)", "Unique, Nullable", "None", "Corporate unique employee ID number."],
                ["specialization", "VARCHAR(200)", "Required", "''", "Primary area of expertise.\nDropdown Options:\n- 'System Architecture'\n- 'Database Optimization'\n- 'Cloud Infrastructure'\n- 'Security & Compliance'\n- 'Performance Tuning'\n- 'Python/Backend'\n- 'React/Frontend'\n- 'DevOps/Infrastructure'\n- 'QA/Testing'\n- 'Database/SQL'"],
                ["phone", "VARCHAR(20)", "Required", "''", "Contact telephone number."],
                ["hire_date", "DATETIME", "Nullable", "None", "Date of joining the company."],
                ["status", "VARCHAR(20)", "Required", "'Active'", "Account status.\nDropdown Options:\n- 'Active'\n- 'Inactive'\n- 'Suspended'"],
                ["is_active", "BOOLEAN", "Required", "True", "Controls if the user can log in."],
                ["created_at", "DATETIME", "Required", "func.now()", "Timestamp of account creation."]
            ]
        },
        # --- roles ---
        {
            "table": "roles",
            "fields": [
                ["id", "INTEGER", "Primary Key, Auto-increment", "None", "Unique role ID."],
                ["name", "VARCHAR(100)", "Required, Unique", "None", "Role name label."],
                ["description", "VARCHAR(500)", "Required", "''", "Short description of role privileges."],
                ["permissions", "JSON", "Required", "[]", "List of allowed permissions strings."],
                ["created_at", "DATETIME", "Required", "func.now()", "Timestamp of role creation."]
            ]
        },
        # --- groups ---
        {
            "table": "groups",
            "fields": [
                ["id", "INTEGER", "Primary Key, Auto-increment", "None", "Unique assignment group ID."],
                ["name", "VARCHAR(100)", "Required, Unique", "None", "Assignment group name."],
                ["description", "VARCHAR(500)", "Required", "''", "Group description / functions."],
                ["user_ids", "JSON", "Required", "[]", "List of user IDs belonging to this group."],
                ["member_count", "INTEGER", "Required", "0", "Total number of users in this group."],
                ["created_at", "DATETIME", "Required", "func.now()", "Timestamp of group creation."]
            ]
        },
        # --- licenses ---
        {
            "table": "licenses",
            "fields": [
                ["id", "INTEGER", "Primary Key, Auto-increment", "None", "Unique license type ID."],
                ["name", "VARCHAR(100)", "Required, Unique", "None", "License tier name."],
                ["description", "VARCHAR(500)", "Required", "''", "Description of license terms."],
                ["max_users", "INTEGER", "Required", "None", "Maximum number of users allowed under this tier."],
                ["features", "JSON", "Required", "[]", "List of unlocked platform module codes."],
                ["created_at", "DATETIME", "Required", "func.now()", "Timestamp of license registration."]
            ]
        },
        # --- workflows ---
        {
            "table": "workflows",
            "fields": [
                ["id", "INTEGER", "Primary Key, Auto-increment", "None", "Workflow ID for low-code builder."],
                ["name", "VARCHAR(200)", "Required, Unique", "None", "Name of the workflow template."],
                ["description", "TEXT", "Nullable", "None", "Detailed description of the workflow purpose."],
                ["created_at", "DATETIME", "Required", "func.now()", "Timestamp of workflow creation."]
            ]
        },
        # --- stages ---
        {
            "table": "stages",
            "fields": [
                ["id", "INTEGER", "Primary Key, Auto-increment", "None", "Stage unique identifier."],
                ["workflow_id", "INTEGER", "Foreign Key (workflows.id)", "None", "ID of the parent workflow template."],
                ["name", "VARCHAR(200)", "Required", "None", "Name of the workflow step/stage."],
                ["order", "INTEGER", "Required", "1", "Sorting order sequence in the workflow progress bar."]
            ]
        },
        # --- task_definitions ---
        {
            "table": "task_definitions",
            "fields": [
                ["id", "INTEGER", "Primary Key, Auto-increment", "None", "Task identifier."],
                ["stage_id", "INTEGER", "Foreign Key (stages.id)", "None", "ID of the parent workflow stage."],
                ["name", "VARCHAR(200)", "Required", "None", "Task name or checklist instruction description."],
                ["task_type", "VARCHAR(100)", "Required", "None", "Technical category of task (e.g. checkbox, API, form)."],
                ["config", "JSON", "Required", "{}", "JSON object storing custom configuration rules."],
                ["created_at", "DATETIME", "Required", "func.now()", "Timestamp of task template creation."]
            ]
        },
        # --- knowledge_articles ---
        {
            "table": "knowledge_articles",
            "fields": [
                ["id", "INTEGER", "Primary Key, Auto-increment", "None", "Unique knowledge base article ID."],
                ["title", "VARCHAR(200)", "Required", "None", "Title/topic of the article."],
                ["content", "TEXT", "Required", "None", "Full HTML or plain text body of the resolution guide."],
                ["tags", "JSON", "Required", "[]", "List of filter tags attached to the article."],
                ["created_at", "DATETIME", "Required", "func.now()", "Timestamp of publication."]
            ]
        },
        # --- FRONTEND ONLY FORM FIELDS ---
        {
            "table": "Frontend Form (Incidents UI)",
            "fields": [
                ["caller", "Dropdown (String)", "Form-Only (Datalist)", "None", "Customer Name.\nOptions:\n- Acme Rail Systems\n- MetroLine Services\n- Urban Transit Corp"],
                ["customer_contract", "Dropdown (String)", "Form-Only (Select)", "None", "Active service level contract code.\nOptions:\n- CTR-2024-001\n- CTR-2024-014\n- CTR-2023-220"],
                ["requestor_name", "Dropdown (String)", "Form-Only (Datalist)", "None", "Name of the customer personnel reporting issue. Choices filter dynamically based on Caller."],
                ["requestor_contact", "Dropdown (String)", "Form-Only (Datalist)", "None", "Phone number of the requestor. Filters dynamically based on Caller."],
                ["srlm_system", "Dropdown (String)", "Form-Only (Select)", "None", "Base system category.\nOptions:\n- 'LM' (Loitering Munition)\n- 'GCS' (Ground Control System)\n- 'TMV' (Tactical Mobility Vehicle)\n- 'Simulator' (Flight Simulator)"],
                ["platform_variant", "Dropdown (String)", "Form-Only (Select)", "None", "Unit model. Filters dynamically based on srlm_system:\n- 'LM Alpha - (LM-0001)'\n- 'LM Bravo - (LM-0002)'\n- 'GCS North - (GCS-0101)'\n- 'TMV Ranger - (TMV-2201)'\n- 'Simulator Prime - (SIM-3001)'"],
                ["sub_system", "Dropdown (String)", "Form-Only (Select)", "None", "System module section. Filters dynamically based on srlm_system:\n- Airframe, Propulsion, Avionics, Comm, Payload (for LM)\n- Computing, Comm, Display, Control Interface, Power (for GCS)\n- Powertrain, Electrical, Comm, Navigation, Payload Handling (for TMV)\n- Computing, Display, Control, Audio, Software (for Simulator)"],
                ["line_replaceable_unit", "Dropdown (String)", "Form-Only (Select)", "None", "Part to replace (LRU). Filters dynamically based on selected Sub-system & Platform."],
                ["warranty_status", "Text", "Form-Only (Display)", "None", "Displays warranty eligibility status of the drone (e.g. Active, Expired)."],
                ["last_serviced_on", "Text", "Form-Only (Display)", "None", "Displays the date of last maintenance action."]
            ]
        },
        # --- LRU FORM (ADMIN UI) ---
        {
            "table": "LRU Form (Admin UI)",
            "fields": [
                ["lru_serial_number", "String", "Required", "None", "Unique serial number of the LRU part."],
                ["lru_name", "String", "Required", "None", "Descriptive component part name."],
                ["sap_part_number", "String", "Required (14 digits)", "None", "14 digit part number generated on SAP."],
                ["customer_part_number", "String", "Required (8 digits)", "None", "08 digit part number generated for the customer."],
                ["make", "String", "Nullable", "None", "Manufacturer brand of the LRU part."],
                ["model", "String", "Nullable", "None", "Model details of the LRU part."],
                ["software_version", "String", "Nullable", "None", "Installed firmware/software version on the LRU."],
                ["platform_variant", "Dropdown (String)", "Required", "None", "Links LRU to a parent platform variant model.\nOptions:\n- LM Alpha - (LM-0001)\n- LM Bravo - (LM-0002)\n- GCS North - (GCS-0101)\n- TMV Ranger - (TMV-2201)\n- Simulator Prime - (SIM-3001)"],
                ["sub_system", "String", "Required", "None", "Associated structural sub-system (e.g. Propulsion, Computing, Navigation)."]
            ]
        },
        # --- TMV FORM (ADMIN UI) ---
        {
            "table": "TMV Form (Admin UI)",
            "fields": [
                ["sap_part_number", "String", "Required (14 digits)", "None", "14 digit part number generated on SAP."],
                ["customer_part_number", "String", "Required (8 digits)", "None", "08 digit part number generated for customer."],
                ["serial_number", "String", "Required", "None", "Unique hardware chassis/VIN code of the TMV."],
                ["unit_name", "String", "Required", "None", "Descriptive unit name (e.g. TMV Ranger)."],
                ["srlm_system", "Dropdown (String)", "Read-only (Default: 'TMV')", "TMV", "System class category."],
                ["customer_id", "Dropdown (String)", "Required", "None", "Owner customer account selection."],
                ["contract_id", "Dropdown (String)", "Required", "None", "Associated customer deployment contract."],
                ["warranty_valid_from", "Date / String", "Required", "None", "Warranty start date (YYYY-MM-DD)."],
                ["warranty_valid_to", "Date / String", "Required", "None", "Warranty end date (YYYY-MM-DD)."],
                ["make", "String", "Nullable", "None", "Vehicle manufacturer brand."],
                ["model", "String", "Nullable", "None", "Vehicle model type details."]
            ]
        }
    ]

    current_row = 5
    for table_group in schema_data:
        t_name = table_group["table"]
        fields = table_group["fields"]
        
        # Write rows
        for f in fields:
            ws.row_dimensions[current_row].height = 65 if "\n" in f[4] else 20
            ws.cell(row=current_row, column=1, value=t_name).font = table_font
            ws.cell(row=current_row, column=1).fill = table_fill
            ws.cell(row=current_row, column=1).alignment = align_left
            ws.cell(row=current_row, column=1).border = thin_border
            
            for col_idx, val in enumerate(f, 2):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = align_left
                
            current_row += 1
            
    # Auto-adjust widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col[3:]:
            if cell.value is not None:
                # Get max line length
                lines = str(cell.value).split("\n")
                for line in lines:
                    max_len = max(max_len, len(line))
        
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # Cap column F (dropdown options) to a readable width like 50 and set wrap text
    ws.column_dimensions['F'].width = 60

    filename = "Vajra_Database_Schema.xlsx"
    wb.save(filename)
    print(f"[SUCCESS] Schema generated at: {filename}")

if __name__ == "__main__":
    generate_schema_excel()
